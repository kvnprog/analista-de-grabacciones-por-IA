from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from fastapi import UploadFile
import io

# Colores de referencia para resaltar emociones negativas/positivas de un vistazo
EMOCIONES_NEGATIVAS = {"frustrado", "molesto", "enojado", "irritado", "confundido", "insatisfecho"}
EMOCIONES_POSITIVAS = {"satisfecho", "amable", "empatico", "empático", "contento", "feliz"}

FILL_NEGATIVO = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_POSITIVO = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_NEUTRAL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")


def _fill_por_emocion(emocion: str) -> PatternFill | None:
    if not emocion:
        return None
    emocion_lower = emocion.strip().lower()
    if emocion_lower in EMOCIONES_NEGATIVAS:
        return FILL_NEGATIVO
    if emocion_lower in EMOCIONES_POSITIVAS:
        return FILL_POSITIVO
    if emocion_lower == "neutral":
        return FILL_NEUTRAL
    return None


def _estilizar_encabezado(ws, fila=1):
    for cell in ws[fila]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def read_questions_from_excel(file: UploadFile) -> list[str]:
    """
    Lee un Excel donde la primera fila contiene las preguntas, una por columna.
    Regresa una lista de strings con las preguntas encontradas, en el mismo orden
    en que aparecen en el archivo.
    """
    file.file.seek(0)
    content = file.file.read()

    if not content or len(content) == 0:
        raise Exception("El archivo de preguntas está vacío")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise Exception("No se pudo leer el archivo Excel. Verifica que sea un .xlsx válido")

    ws = wb.active
    preguntas = []

    primera_fila = next(ws.iter_rows(min_row=1, max_row=1), None)
    if primera_fila:
        for cell in primera_fila:
            if cell.value is not None and str(cell.value).strip():
                preguntas.append(str(cell.value).strip())

    wb.close()

    if not preguntas:
        raise Exception("No se encontraron preguntas en la primera fila del Excel")

    return preguntas


def create_analysis_excel(analysis: dict) -> str:
    wb = Workbook()

    # Detectamos qué formas de búsqueda se usaron, revisando todos los items
    todas_las_palabras = set()
    lista_preguntas_excel = []
    tiene_busqueda_texto = False
    tiene_emociones = False

    for item in analysis:
        todas_las_palabras.update(item.get('palabras', {}).keys())

        if item.get('preguntas_excel') and not lista_preguntas_excel:
            lista_preguntas_excel = list(item['preguntas_excel'].keys())

        if item.get('busqueda', {}).get('texto'):
            tiene_busqueda_texto = True

        if item.get('emociones'):
            tiene_emociones = True

    lista_palabras = sorted(list(todas_las_palabras))
    tiene_palabras = len(lista_palabras) > 0
    tiene_preguntas_excel = len(lista_preguntas_excel) > 0

    # ============================================================
    # HOJA 1: RESUMEN — vista general de todos los archivos procesados
    # ============================================================
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Archivo", "Estado", "Formas de búsqueda aplicadas"])
    _estilizar_encabezado(ws_resumen)

    for item in analysis:
        formas = []
        if item.get('palabras'):
            formas.append("Palabras clave")
        if item.get('preguntas_excel'):
            formas.append("Preguntas Excel")
        if item.get('busqueda', {}).get('texto'):
            formas.append("Búsqueda libre")
        if item.get('emociones'):
            formas.append("Emociones")

        estado = "Con error" if item.get('error') else "Procesado"

        ws_resumen.append([
            item.get('audio', 'N/A'),
            estado,
            ", ".join(formas) if formas else "Ninguna"
        ])

        if item.get('error'):
            fila_idx = ws_resumen.max_row
            for col in range(1, 4):
                ws_resumen.cell(row=fila_idx, column=col).fill = FILL_NEGATIVO

    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 40

    # ============================================================
    # HOJA: PALABRAS CLAVE — una columna por palabra buscada
    # ============================================================
    if tiene_palabras:
        ws_palabras = wb.create_sheet(title="Palabras Clave")
        ws_palabras.append(["Archivo"] + lista_palabras)
        _estilizar_encabezado(ws_palabras)

        for item in analysis:
            conteo = item.get('palabras', {})
            fila = [item.get('audio', 'N/A')] + [conteo.get(p, 0) for p in lista_palabras]
            ws_palabras.append(fila)

        ws_palabras.column_dimensions['A'].width = 30
        for i in range(2, len(lista_palabras) + 2):
            ws_palabras.column_dimensions[get_column_letter(i)].width = 15

        # Gráfico de totales de palabras
        ws_charts = wb.create_sheet(title="Gráfico Palabras")
        ws_charts.append(["Palabra", "Total"])

        totales = {}
        for item in analysis:
            for palabra, cantidad in item.get('palabras', {}).items():
                totales[palabra] = totales.get(palabra, 0) + cantidad

        for i, (palabra, total) in enumerate(totales.items(), start=2):
            ws_charts.cell(row=i, column=1, value=palabra)
            ws_charts.cell(row=i, column=2, value=total)

        if totales:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Total de Palabras Clave"
            chart.y_axis.title = 'Cantidad'
            chart.x_axis.title = 'Palabras'
            chart.legend = None

            data = Reference(ws_charts, min_col=2, min_row=1, max_row=len(totales) + 1, max_col=2)
            cats = Reference(ws_charts, min_col=1, min_row=2, max_row=len(totales) + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_charts.add_chart(chart, "D2")

    # ============================================================
    # HOJA: PREGUNTAS EXCEL — una columna por pregunta del Excel subido
    # ============================================================
    if tiene_preguntas_excel:
        ws_preguntas = wb.create_sheet(title="Preguntas Excel")
        ws_preguntas.append(["Archivo"] + lista_preguntas_excel)
        _estilizar_encabezado(ws_preguntas)

        for item in analysis:
            respuestas = item.get('preguntas_excel', {}) or {}
            fila = [item.get('audio', 'N/A')] + [respuestas.get(p, 'N/A') for p in lista_preguntas_excel]
            ws_preguntas.append(fila)

        ws_preguntas.column_dimensions['A'].width = 30
        for i in range(2, len(lista_preguntas_excel) + 2):
            ws_preguntas.column_dimensions[get_column_letter(i)].width = 40

    # ============================================================
    # HOJA: BÚSQUEDA LIBRE — la pregunta abierta que escribió el usuario
    # ============================================================
    if tiene_busqueda_texto:
        ws_busqueda = wb.create_sheet(title="Búsqueda Libre")
        ws_busqueda.append(["Archivo", "Pregunta Realizada", "Resultado"])
        _estilizar_encabezado(ws_busqueda)

        for item in analysis:
            busqueda = item.get('busqueda', {})
            ws_busqueda.append([
                item.get('audio', 'N/A'),
                busqueda.get('texto', 'Sin consulta'),
                busqueda.get('detalle', 'N/A')
            ])

        ws_busqueda.column_dimensions['A'].width = 30
        ws_busqueda.column_dimensions['B'].width = 35
        ws_busqueda.column_dimensions['C'].width = 60

    # ============================================================
    # HOJA: EMOCIONES — bot vs cliente, con colores por tipo de emoción
    # ============================================================
    if tiene_emociones:
        ws_emociones = wb.create_sheet(title="Emociones")
        ws_emociones.append([
            "Archivo",
            "Emoción Bot/Agente", "Detalle Bot/Agente",
            "Emoción Cliente", "Detalle Cliente",
            "Confianza"
        ])
        _estilizar_encabezado(ws_emociones)

        for item in analysis:
            emociones = item.get('emociones', {}) or {}
            emocion_bot = emociones.get('bot', '')
            emocion_cliente = emociones.get('cliente', '')

            ws_emociones.append([
                item.get('audio', 'N/A'),
                emocion_bot or 'N/A',
                emociones.get('bot_detalle', '') or 'N/A',
                emocion_cliente or 'N/A',
                emociones.get('cliente_detalle', '') or 'N/A',
                emociones.get('confianza', '') or 'N/A'
            ])

            fila_idx = ws_emociones.max_row

            fill_bot = _fill_por_emocion(emocion_bot)
            if fill_bot:
                ws_emociones.cell(row=fila_idx, column=2).fill = fill_bot

            fill_cliente = _fill_por_emocion(emocion_cliente)
            if fill_cliente:
                ws_emociones.cell(row=fila_idx, column=4).fill = fill_cliente

        ws_emociones.column_dimensions['A'].width = 30
        ws_emociones.column_dimensions['B'].width = 18
        ws_emociones.column_dimensions['C'].width = 40
        ws_emociones.column_dimensions['D'].width = 18
        ws_emociones.column_dimensions['E'].width = 40
        ws_emociones.column_dimensions['F'].width = 12

    file_path = "analysis_result.xlsx"
    wb.save(file_path)

    return file_path